import pandas as pd
import requests
import time
import random
from io import StringIO
from datetime import datetime
import concurrent.futures
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ==========================================
# ÏÑ§Ï†ï
# ==========================================
INPUT_CSV = "260130_Earnings.csv"  # ÏûÖÎ†• CSV
OUTPUT_FILE = "Ìï¥Ïô∏Îπ®Í∞ÑÏ§Ñ_ÏôÑÏÑ±.xlsx"  # ÏµúÏ¢Ö Ï∂úÎ†•
OUTPUT_FAILED = "failed_tickers.xlsx"  # Ïã§Ìå® Î™©Î°ù

NUM_QUARTERS = 17
MAX_WORKERS = 4

# ÏµúÏã† Ïã§Ï†Å Í∏∞Ï§Ä (Ïù¥Í≤ÉÎ≥¥Îã§ Ïò§ÎûòÎêòÎ©¥ Îß® ÏïÑÎûòÎ°ú)
# Oct 2025 Ïù¥ÏÉÅÎßå Ï†ïÏÉÅ (Q4 2025 = Oct~Dec)
MIN_DATE = (2025, 10)  # (ÎÖÑ, Ïõî)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept-Language': 'en-US,en;q=0.9',
}

QUARTER_LABELS = ['4Q21', '1Q22', '2Q22', '3Q22', '4Q22', '1Q23', '2Q23', '3Q23', '4Q23', '1Q24', '2Q24', '3Q24', '4Q24', '1Q25', '2Q25', '3Q25', '4Q25']

# Í±∞ÎûòÏÜå ‚Üí Íµ≠Í∞Ä Îß§Ìïë
EXCHANGE_TO_COUNTRY = {
    'NMS': 'United States', 'NYQ': 'United States', 'ASE': 'United States', 
    'PCX': 'United States', 'NGM': 'United States', 'NAS': 'United States',
    'STO': 'Sweden', 'PAR': 'France', 'AMS': 'Netherlands', 'BRU': 'Belgium', 
    'LSE': 'United Kingdom', 'FRA': 'Germany', 'SWX': 'Switzerland',
    'JPX': 'Japan', 'TYO': 'Japan',
    'HKG': 'Hong Kong',
    'KSC': 'South Korea', 'KOE': 'South Korea',
    'TAI': 'Taiwan', 'SHH': 'China', 'SHZ': 'China',
    'IDX': 'Indonesia', 'NSI': 'India', 'BOM': 'India',
    'BMV': 'Mexico', 'SAO': 'Brazil', 'JNB': 'South Africa', 'TAD': 'Israel',
}

# ==========================================
# Ìó¨Ìçº Ìï®Ïàò
# ==========================================
def parse_money_string(value_str):
    if not isinstance(value_str, str):
        return value_str
    s = value_str.strip().replace(',', '')
    if s == '-':
        return 0
    try:
        if s.endswith('B'):
            return float(s[:-1]) * 1_000_000_000
        elif s.endswith('M'):
            return float(s[:-1]) * 1_000_000
        elif s.endswith('K'):
            return float(s[:-1]) * 1_000
        elif s.endswith('%'):
            return float(s[:-1])
        else:
            return float(s)
    except:
        return 0

def parse_date_to_year_month(date_str):
    """
    Îã§ÏñëÌïú ÎÇ†Ïßú ÌòïÏãùÏùÑ (ÎÖÑ, Ïõî)Î°ú ÌååÏã±
    Q4 2025 ‚Üí (2025, 12)
    Jun 2026 ‚Üí (2026, 6)
    Dec 31, 2025 ‚Üí (2025, 12)
    Dec'25 ‚Üí (2025, 12)
    """
    if not isinstance(date_str, str):
        return None, None
    
    import re
    date_str = date_str.strip()
    
    month_map = {
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
        'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
    }
    
    quarter_to_month = {'1': 3, '2': 6, '3': 9, '4': 12}
    
    date_lower = date_str.lower()
    
    # "Q4 2025" ÌòïÏãù
    match = re.search(r'q(\d)\s*(\d{4})', date_lower)
    if match:
        q = match.group(1)
        year = int(match.group(2))
        return year, quarter_to_month.get(q, 12)
    
    # "Jun 2026" ÎòêÎäî "June 2026" ÌòïÏãù
    for month_name, month_num in month_map.items():
        if month_name in date_lower:
            match = re.search(r'(\d{4})', date_str)
            if match:
                return int(match.group(1)), month_num
            # '25, '26 ÌòïÏãù
            match = re.search(r"'(\d{2})", date_str)
            if match:
                return 2000 + int(match.group(1)), month_num
    
    # "Dec 31, 2025" ÌòïÏãù
    match = re.search(r'(\d{4})', date_str)
    if match:
        year = int(match.group(1))
        for month_name, month_num in month_map.items():
            if month_name in date_lower:
                return year, month_num
    
    return None, None

def is_recent_enough(date_str):
    """ÏµúÏã† Ïã§Ï†ÅÏù∏ÏßÄ ÌôïÏù∏ (MIN_DATE Í∏∞Ï§Ä)"""
    year, month = parse_date_to_year_month(date_str)
    if year is None or month is None:
        return False
    
    min_year, min_month = MIN_DATE
    
    if year > min_year:
        return True
    elif year == min_year and month >= min_month:
        return True
    return False

def convert_date_format(date_str):
    """
    Îã§ÏñëÌïú ÎÇ†Ïßú ÌòïÏãùÏùÑ ÌÜµÏùºÎêú ÌòïÏãùÏúºÎ°ú Î≥ÄÌôò
    Q4 2025 ‚Üí Dec'25
    Jun 2026 ‚Üí Jun'26
    Dec 31, 2025 ‚Üí Dec'25
    """
    if not isinstance(date_str, str):
        return str(date_str)
    
    import re
    
    month_map = {
        1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
        7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
    }
    
    year, month = parse_date_to_year_month(date_str)
    if year and month:
        year_short = str(year)[2:]
        return f"{month_map[month]}'{year_short}"
    
    return date_str

def map_quarters_from_latest(latest_date_raw, num_values):
    """
    ÏµúÏã† Îç∞Ïù¥ÌÑ∞Î•º 4Q25Ïóê, Í∑∏ Ïù¥Ï†ÑÏùÑ 3Q25, 2Q25... ÏàúÏÑúÎ°ú Î∞∞Ïπò
    ÌöåÍ≥ÑÏó∞ÎèÑ ÏÉÅÍ¥ÄÏóÜÏù¥ ÏÉÅÎåÄÏ†Å ÏúÑÏπòÎ°ú ÌÜµÏùº
    """
    labels = []
    for i in range(num_values):
        # 4Q25Í∞Ä index 16 (ÎßàÏßÄÎßâ), Í±∞Í∏∞ÏÑú Ïó≠ÏàúÏúºÎ°ú
        target_idx = 16 - i
        if target_idx >= 0 and target_idx < len(QUARTER_LABELS):
            labels.append(QUARTER_LABELS[target_idx])
        else:
            labels.append(None)
    return labels

def get_industry(ticker):
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        return info.get('industry', info.get('sector', 'N/A'))
    except:
        return 'N/A'

def get_stock_analysis_exchange(ticker, company_name=None):
    exchange_map = {
        'NMS': None, 'NYQ': None, 'ASE': None, 'PCX': None, 'NGM': None, 'NAS': None,
        'STO': 'sto', 'PAR': 'epa', 'AMS': 'ams', 'BRU': 'bru', 'LSE': 'lse', 'FRA': 'fra', 'SWX': 'swx',
        'JPX': 'tyo', 'TYO': 'tyo', 'HKG': 'hkg', 'KSC': 'ksc', 'KOE': 'koe',
        'TAI': 'tai', 'SHH': 'shh', 'SHZ': 'shz', 'IDX': 'idx', 'NSI': 'nse', 'BOM': 'bse',
        'BMV': 'bmv', 'SAO': 'sao', 'JNB': 'jse', 'TAD': 'tase',
    }
    
    search_attempts = [ticker]
    
    if ticker.replace(' ', '').replace('-', '').isdigit():
        ticker_clean = ticker.replace(' ', '')
        if len(ticker_clean) <= 4:
            search_attempts.extend([f"{ticker_clean}.T", f"{ticker_clean}.HK"])
        else:
            search_attempts.extend([f"{ticker_clean}.HK", f"{ticker_clean}.NS", f"{ticker_clean}.BO", f"{ticker_clean}.SR"])
    elif ' ' in ticker:
        ticker_hyphen = ticker.replace(' ', '-')
        search_attempts.extend([f"{ticker_hyphen}.ST", f"{ticker}.ST"])
    elif ticker.isupper() and len(ticker) > 4:
        search_attempts.extend([f"{ticker}.NS", f"{ticker}.BO"])
    else:
        search_attempts.extend([f"{ticker}.ST", f"{ticker}.PA", f"{ticker}.L", f"{ticker}.T", f"{ticker}.HK"])
    
    for attempt_ticker in search_attempts:
        try:
            stock = yf.Ticker(attempt_ticker)
            info = stock.info
            symbol = info.get('symbol')
            exchange = info.get('exchange', '')
            if symbol and exchange:
                sa_code = exchange_map.get(exchange)
                country = EXCHANGE_TO_COUNTRY.get(exchange, 'Unknown')
                return sa_code, exchange, attempt_ticker, country
        except:
            continue
    return None, None, ticker, 'United States'

# ==========================================
# ÌÅ¨Î°§ÎßÅ Ìï®Ïàò
# ==========================================
def process_ticker(ticker_data):
    if isinstance(ticker_data, tuple):
        raw_ticker, company_name = ticker_data
    else:
        raw_ticker = ticker_data
        company_name = None
    
    ticker = str(raw_ticker).strip().replace('.', '-').replace(' ', '-').lower()
    sa_exchange, original_exchange, matched_ticker, country = get_stock_analysis_exchange(raw_ticker, company_name)
    
    # ÌïúÍµ≠ Ï£ºÏãù Ï†úÏô∏
    if sa_exchange in ['ksc', 'koe']:
        return {'status': 'failed', 'ticker': raw_ticker, 'reason': 'Korean Stock (Skipped)'}
    
    if isinstance(raw_ticker, str) and raw_ticker.startswith('A') and len(raw_ticker) == 7 and raw_ticker[1:].isdigit():
        return {'status': 'failed', 'ticker': raw_ticker, 'reason': 'Korean Stock Code (Skipped)'}
    
    if sa_exchange not in ['tyo', 'hkg', None]:
        if any(char.isdigit() for char in ticker) and not ticker.isalpha():
            return {'status': 'failed', 'ticker': raw_ticker, 'reason': f'Non-supported ticker'}

    if sa_exchange:
        url = f"https://stockanalysis.com/quote/{sa_exchange}/{raw_ticker.upper().replace(' ', '-')}/financials/?p=quarterly"
    else:
        url = f"https://stockanalysis.com/stocks/{ticker}/financials/?p=quarterly"
    
    retry_count = 0
    time.sleep(random.uniform(1.0, 3.0))

    while retry_count < 3:
        try:
            response = requests.get(url, headers=HEADERS, timeout=10)
            if response.status_code == 200:
                break
            elif response.status_code == 404:
                return {'status': 'failed', 'ticker': raw_ticker, 'reason': f'404 Not Found'}
            elif response.status_code == 429:
                time.sleep(random.uniform(10, 20))
                retry_count += 1
            else:
                return {'status': 'failed', 'ticker': raw_ticker, 'reason': f'Error {response.status_code}'}
        except:
            retry_count += 1
            time.sleep(2)
            
    if retry_count >= 3:
        return {'status': 'failed', 'ticker': raw_ticker, 'reason': 'Connection Timeout'}

    try:
        dfs = pd.read_html(StringIO(response.text))
        if not dfs:
            return {'status': 'failed', 'ticker': raw_ticker, 'reason': 'No Table Found'}
        df_fin = dfs[0]

        date_cols = df_fin.columns[1:].tolist()
        latest_date_raw = date_cols[0]
        
        # ÌäúÌîåÏù∏ Í≤ΩÏö∞ Ï≤´ Î≤àÏß∏ Í∞í ÏÇ¨Ïö©
        if isinstance(latest_date_raw, tuple):
            latest_date_raw = str(latest_date_raw[0])
        
        # ÏµúÏã† Ïã§Ï†Å Ï≤¥ÌÅ¨ (Ïò§ÎûòÎêòÎ©¥ ÌîåÎûòÍ∑∏Îßå ÏÑ§Ï†ï, Ï†úÏô∏ÌïòÏßÄ ÏïäÏùå)
        is_outdated = not is_recent_enough(latest_date_raw)
        
        # ÎÇ†Ïßú ÌòïÏãù Î≥ÄÌôò (Q4 2025 ‚Üí Dec'25)
        latest_date_str = convert_date_format(latest_date_raw)
        
        # Ïó∞Í∞Ñ Îç∞Ïù¥ÌÑ∞ Ï≤¥ÌÅ¨
        try:
            if len(date_cols) >= 2:
                d1_raw = date_cols[0] if not isinstance(date_cols[0], tuple) else date_cols[0][0]
                d2_raw = date_cols[1] if not isinstance(date_cols[1], tuple) else date_cols[1][0]
                d1 = pd.to_datetime(str(d1_raw), format='mixed')
                d2 = pd.to_datetime(str(d2_raw), format='mixed')
                if abs((d1 - d2).days) > 250:
                    return {'status': 'failed', 'ticker': raw_ticker, 'reason': 'Annual Data (FY)'}
        except:
            pass

        # Revenue Ï∞æÍ∏∞
        revenue_target = ["Revenue", "Total Revenue", "Net Revenue", "Sales"]
        revenue_row = pd.DataFrame()
        for metric in revenue_target:
            temp = df_fin[df_fin.iloc[:, 0].str.strip().str.lower() == metric.lower()]
            if not temp.empty:
                revenue_row = temp
                break
        
        if revenue_row.empty:
            revenue_values = None
            revenue_growth = 0
        else:
            revenue_values = [parse_money_string(v) for v in revenue_row.iloc[0, 1:].tolist()]
            revenue_values = revenue_values[:NUM_QUARTERS]
            if len(revenue_values) >= 5:
                recent_avg = sum(revenue_values[0:4]) / 4
                past_avg = sum(revenue_values[1:5]) / 4
                revenue_growth = (recent_avg / past_avg) - 1 if past_avg != 0 else 0
            else:
                revenue_growth = 0

        # Operating Income Ï∞æÍ∏∞
        op_target = ["Operating Income", "Operating Profit", "Pretax Income", "Net Income"]
        op_row = pd.DataFrame()
        for metric in op_target:
            temp = df_fin[df_fin.iloc[:, 0].str.contains(metric, case=False, na=False)]
            if not temp.empty:
                op_row = temp
                break
        
        if op_row.empty:
            op_values = None
            op_growth = 0
        else:
            op_values = [parse_money_string(v) for v in op_row.iloc[0, 1:].tolist()]
            op_values = op_values[:NUM_QUARTERS]
            if len(op_values) >= 5:
                recent_avg = sum(op_values[0:4]) / 4
                past_avg = sum(op_values[1:5]) / 4
                op_growth = (recent_avg / past_avg) - 1 if past_avg != 0 else 0
            else:
                op_growth = 0

        # Îëò Îã§ ÏóÜÏúºÎ©¥ Ïã§Ìå®
        if revenue_values is None and op_values is None:
            return {'status': 'failed', 'ticker': raw_ticker, 'reason': 'No Data Found'}
        
        # ÏµúÏÜå 4Î∂ÑÍ∏∞
        if revenue_values and len(revenue_values) < 4:
            revenue_values = None
        if op_values and len(op_values) < 4:
            op_values = None
            
        if revenue_values is None and op_values is None:
            return {'status': 'failed', 'ticker': raw_ticker, 'reason': 'Data < 4 quarters'}

        num_quarters = max(len(revenue_values) if revenue_values else 0, len(op_values) if op_values else 0)
        
        result = {
            'status': 'success',
            'ticker': raw_ticker.upper(),
            'country': country,
            'industry': 'PENDING',
            'latest_date': latest_date_str,
            'latest_date_raw': latest_date_raw,  # ÏõêÎ≥∏ Ï†ÄÏû• (Q4 2025 ÌòïÏãù)
            'is_outdated': is_outdated,  # Ïò§ÎûòÎêú Ïã§Ï†Å Ïó¨Î∂Ä
            'revenue_values': revenue_values,
            'revenue_growth': revenue_growth * 100,
            'op_values': op_values,
            'op_growth': op_growth * 100,
            'num_quarters': num_quarters
        }
        
        status_icon = "‚ö†Ô∏è Íµ¨Ïã§Ï†Å" if is_outdated else "‚úÖ"
        print(f"[{raw_ticker.upper()}] {status_icon} {country} - {num_quarters}Î∂ÑÍ∏∞ ({latest_date_str})")
        return result

    except Exception as e:
        return {'status': 'failed', 'ticker': raw_ticker, 'reason': str(e)[:50]}

# ==========================================
# ÏóëÏÖÄ ÏÉùÏÑ± Ìï®ÏàòÎì§
# ==========================================
def sort_dataframe(df):
    df = df.copy()
    
    def country_priority(c):
        c_str = str(c).strip().lower() if pd.notna(c) else ''
        if 'united states' in c_str:
            return 0
        elif 'japan' in c_str:
            return 1
        else:
            return 2
    
    # is_outdated Ïª¨ÎüºÏù¥ ÏûàÏúºÎ©¥ ÏÇ¨Ïö©, ÏóÜÏúºÎ©¥ ÎÇ†ÏßúÎ°ú ÌåêÎã®
    if 'is_outdated' in df.columns:
        df['_is_outdated'] = df['is_outdated'].fillna(False)
    else:
        df['_is_outdated'] = ~df['Latest_Date'].apply(is_recent_enough)
    
    df['_country_order'] = df['Country'].apply(country_priority)
    df['_growth'] = pd.to_numeric(df['Growth_Rate'], errors='coerce').fillna(0)
    
    # Ï†ïÎ†¨: outdated Îß® ÏïÑÎûò ‚Üí Country ‚Üí Growth
    df = df.sort_values(
        by=['_is_outdated', '_country_order', '_growth'],
        ascending=[True, True, False]  # outdated=FalseÍ∞Ä Î®ºÏ†Ä (ÏúÑÎ°ú)
    )
    
    df = df.drop(columns=['_country_order', '_is_outdated', '_growth'], errors='ignore')
    return df.reset_index(drop=True)

def find_outdated_start(df):
    """outdated Îç∞Ïù¥ÌÑ∞ ÏãúÏûë ÏúÑÏπò (Ïà®ÍπÄÏö©)"""
    if 'is_outdated' not in df.columns:
        return None
    for i, row in df.iterrows():
        if row.get('is_outdated', False):
            return i + 3  # Ìó§ÎçîÍ∞Ä 2Ìñâ
    return None

def calc_rolling_growth(df, quarter_cols):
    """Î°§ÎßÅ 4Î∂ÑÍ∏∞ ÏÑ±Ïû•Î•† Í≥ÑÏÇ∞"""
    results = []
    # 4Q25, 3Q25, 2Q25, 1Q25 ÏúÑÏπò: index 16, 15, 14, 13
    target_indices = [13, 14, 15, 16]
    
    for row_idx, row in df.iterrows():
        row_results = []
        for pos in target_indices:
            try:
                if pos >= len(quarter_cols) or pos < 4:
                    row_results.append(None)
                    continue
                
                # Ìï¥Îãπ Î∂ÑÍ∏∞ÏôÄ Ïù¥Ï†Ñ 7Í∞ú Î∂ÑÍ∏∞ Îç∞Ïù¥ÌÑ∞ ÌôïÏù∏
                recent_vals = []
                prev_vals = []
                for i in range(4):
                    q_idx = pos - i
                    if q_idx >= 0 and q_idx < len(quarter_cols):
                        val = row.get(quarter_cols[q_idx])
                        if pd.notna(val) and val != '':
                            recent_vals.append(float(val))
                    
                    p_idx = pos - 1 - i
                    if p_idx >= 0 and p_idx < len(quarter_cols):
                        val = row.get(quarter_cols[p_idx])
                        if pd.notna(val) and val != '':
                            prev_vals.append(float(val))
                
                if len(recent_vals) >= 4 and len(prev_vals) >= 4:
                    recent_avg = sum(recent_vals[:4]) / 4
                    prev_avg = sum(prev_vals[:4]) / 4
                    if prev_avg != 0:
                        growth = (recent_avg / prev_avg) - 1
                        row_results.append(growth)
                    else:
                        row_results.append(None)
                else:
                    row_results.append(None)
            except:
                row_results.append(None)
        results.append(row_results)
    return results

def get_highlight_cells(df, quarter_cols, data_start_row=3, quarter_start_col=7):
    """10% Ïù¥ÏÉÅ ÏÑ±Ïû• ÏÖÄ ÌïòÏù¥ÎùºÏù¥Ìä∏"""
    highlights = []
    for row_idx, row in df.iterrows():
        for pos in range(4, len(quarter_cols)):
            try:
                recent_vals = []
                prev_vals = []
                for i in range(4):
                    val = row.get(quarter_cols[pos-i])
                    if pd.notna(val) and val != '':
                        recent_vals.append(float(val))
                    val2 = row.get(quarter_cols[pos-1-i])
                    if pd.notna(val2) and val2 != '':
                        prev_vals.append(float(val2))
                
                if len(recent_vals) >= 4 and len(prev_vals) >= 4:
                    recent_avg = sum(recent_vals[:4]) / 4
                    prev_avg = sum(prev_vals[:4]) / 4
                    if prev_avg != 0 and recent_avg / prev_avg >= 1.1:
                        excel_row = row_idx + data_start_row
                        excel_col = quarter_start_col + pos
                        highlights.append((excel_row, excel_col))
            except:
                continue
    return highlights

def create_final_excel(success_data, df_earnings, output_file):
    # Îç∞Ïù¥ÌÑ∞ÌîÑÎ†àÏûÑ ÏÉùÏÑ±
    rev_rows = []
    op_rows = []
    
    for data in success_data:
        # ÏµúÏã† Î∂ÑÍ∏∞ Ï†ïÎ≥¥Î°ú Î∂ÑÍ∏∞ ÎùºÎ≤® Í≥ÑÏÇ∞
        latest_raw = data.get('latest_date_raw', data['latest_date'])
        is_outdated = data.get('is_outdated', False)
        
        # Revenue Îç∞Ïù¥ÌÑ∞
        if data['revenue_values']:
            rev_row = {
                'Ticker': data['ticker'],
                'Country': data['country'],
                'Industry': data['industry'],
                'Latest_Date': data['latest_date'],
                'Growth_Rate': data['revenue_growth'],
                'is_outdated': is_outdated
            }
            # Ïã§Ï†ú Î∂ÑÍ∏∞Ïóê ÎßûÍ≤å Î∞∞Ïπò
            quarter_labels = map_quarters_from_latest(latest_raw, len(data['revenue_values']))
            for i, val in enumerate(data['revenue_values']):
                q_label = quarter_labels[i]
                if q_label and q_label in QUARTER_LABELS:
                    rev_row[q_label] = val
            rev_rows.append(rev_row)
        
        # Operating Income Îç∞Ïù¥ÌÑ∞
        if data['op_values']:
            op_row = {
                'Ticker': data['ticker'],
                'Country': data['country'],
                'Industry': data['industry'],
                'Latest_Date': data['latest_date'],
                'Growth_Rate': data['op_growth'],
                'is_outdated': is_outdated
            }
            quarter_labels = map_quarters_from_latest(latest_raw, len(data['op_values']))
            for i, val in enumerate(data['op_values']):
                q_label = quarter_labels[i]
                if q_label and q_label in QUARTER_LABELS:
                    op_row[q_label] = val
            op_rows.append(op_row)
    
    df_rev = pd.DataFrame(rev_rows)
    df_op = pd.DataFrame(op_rows)
    
    # Îπà Ïª¨Îüº Ï±ÑÏö∞Í∏∞
    for q in QUARTER_LABELS:
        if q not in df_rev.columns:
            df_rev[q] = None
        if q not in df_op.columns:
            df_op[q] = None
    
    # Ï†ïÎ†¨
    df_rev_sorted = sort_dataframe(df_rev) if not df_rev.empty else df_rev
    df_op_sorted = sort_dataframe(df_op) if not df_op.empty else df_op
    
    # is_outdated Ïª¨Îüº Ï†úÍ±∞ (ÏóëÏÖÄÏóê ÌëúÏãú ÏïàÌï®, Ï†ïÎ†¨Ïö©ÏúºÎ°úÎßå ÏÇ¨Ïö©)
    if 'is_outdated' in df_rev_sorted.columns:
        df_rev_sorted = df_rev_sorted.drop(columns=['is_outdated'])
    if 'is_outdated' in df_op_sorted.columns:
        df_op_sorted = df_op_sorted.drop(columns=['is_outdated'])
    
    # Î°§ÎßÅ ÏÑ±Ïû•Î•† Î∞è ÌïòÏù¥ÎùºÏù¥Ìä∏ Í≥ÑÏÇ∞
    rev_rolling = calc_rolling_growth(df_rev_sorted, QUARTER_LABELS) if not df_rev_sorted.empty else []
    op_rolling = calc_rolling_growth(df_op_sorted, QUARTER_LABELS) if not df_op_sorted.empty else []
    rev_highlights = get_highlight_cells(df_rev_sorted, QUARTER_LABELS) if not df_rev_sorted.empty else []
    op_highlights = get_highlight_cells(df_op_sorted, QUARTER_LABELS) if not df_op_sorted.empty else []
    
    # ÏõåÌÅ¨Î∂Å ÏÉùÏÑ±
    wb = Workbook()
    
    # Ïä§ÌÉÄÏùº
    pretendard = Font(name='Pretendard', size=10)
    header_font = Font(name='Pretendard', size=10, bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    highlight_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    rolling_header_font = Font(name='Pretendard', size=10, bold=True, color="CC0000")
    rolling_header_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
    
    # ========================================
    # 1. Í∏∞ÏóÖÎ¶¨Ïä§Ìä∏ ÏãúÌä∏
    # ========================================
    ws_list = wb.active
    ws_list.title = "Í∏∞ÏóÖÎ¶¨Ïä§Ìä∏"
    
    ws_list['B2'] = 'Ïã§Ï†Å Í∏∞ÏóÖ Ï†ïÎ≥¥'
    ws_list['B2'].font = Font(name='Pretendard', size=14, bold=True)
    
    list_headers = ['Ticker', 'Company', 'Date', 'Time', 'Quarter Ending', 'Market Cap (mil$)']
    for col, header in enumerate(list_headers, start=2):
        cell = ws_list.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, row in df_earnings.iterrows():
        excel_row = row_idx + 4
        ws_list.cell(row=excel_row, column=2, value=row.get('Ticker', '')).font = pretendard
        ws_list.cell(row=excel_row, column=3, value=row.get('Company', '')).font = pretendard
        ws_list.cell(row=excel_row, column=4, value=row.get('Date', '')).font = pretendard
        ws_list.cell(row=excel_row, column=5, value=row.get('Time', '')).font = pretendard
        ws_list.cell(row=excel_row, column=6, value=row.get('Quarter Ending', '')).font = pretendard
        
        market_cap = row.get(' Market Cap ', row.get('Market Cap', ''))
        if pd.notna(market_cap):
            try:
                mc_str = str(market_cap).replace(',', '').replace(' ', '').strip()
                mc_val = float(mc_str) / 1_000_000
                cell = ws_list.cell(row=excel_row, column=7, value=mc_val)
                cell.number_format = '#,##0'
            except:
                ws_list.cell(row=excel_row, column=7, value=market_cap)
        ws_list.cell(row=excel_row, column=7).font = pretendard
    
    ws_list.column_dimensions['A'].width = 3
    ws_list.column_dimensions['B'].width = 12
    ws_list.column_dimensions['C'].width = 45
    ws_list.column_dimensions['D'].width = 12
    ws_list.column_dimensions['E'].width = 12
    ws_list.column_dimensions['F'].width = 14
    ws_list.column_dimensions['G'].width = 16
    ws_list.auto_filter.ref = f"B3:G{len(df_earnings)+3}"
    
    # ========================================
    # 2. Revenue ÏãúÌä∏ (Î°§ÎßÅ Ï∂îÍ∞Ä)
    # ========================================
    ws_rev = wb.create_sheet("Revenue")
    
    rolling_headers = ['1Q25 Î°§ÎßÅ', '2Q25 Î°§ÎßÅ', '3Q25 Î°§ÎßÅ', '4Q25 Î°§ÎßÅ']
    headers = ['Ticker', 'Country', 'Industry', 'Latest_Date', 'Growth_Rate'] + QUARTER_LABELS + rolling_headers
    
    for col, header in enumerate(headers, start=2):
        cell = ws_rev.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        if header in rolling_headers:
            cell.font = rolling_header_font
            cell.fill = rolling_header_fill
    
    for row_idx, row in df_rev_sorted.iterrows():
        excel_row = row_idx + 3
        ws_rev.cell(row=excel_row, column=2, value=row.get('Ticker', '')).font = pretendard
        ws_rev.cell(row=excel_row, column=3, value=row.get('Country', '')).font = pretendard
        ws_rev.cell(row=excel_row, column=4, value=row.get('Industry', '')).font = pretendard
        ws_rev.cell(row=excel_row, column=5, value=row.get('Latest_Date', '')).font = pretendard
        
        growth_cell = ws_rev.cell(row=excel_row, column=6)
        growth_val = row.get('Growth_Rate', 0)
        if pd.notna(growth_val):
            growth_cell.value = round(float(growth_val), 1)
            growth_cell.number_format = '0.0"%"'
        growth_cell.font = pretendard
        
        for i, q in enumerate(QUARTER_LABELS):
            cell = ws_rev.cell(row=excel_row, column=7+i)
            val = row.get(q)
            if pd.notna(val) and val != '':
                cell.value = val
                cell.number_format = '#,##0'
            cell.font = pretendard
        
        # Î°§ÎßÅ ÏÑ±Ïû•Î•†
        if row_idx < len(rev_rolling):
            for i, rv in enumerate(rev_rolling[row_idx]):
                cell = ws_rev.cell(row=excel_row, column=7+NUM_QUARTERS+i)
                if rv is not None:
                    cell.value = rv
                    cell.number_format = '0.0%'
                    if rv >= 0.1:
                        cell.fill = highlight_fill
                cell.font = pretendard
    
    for (r, c) in rev_highlights:
        ws_rev.cell(row=r, column=c).fill = highlight_fill
    
    ws_rev.column_dimensions['A'].width = 3
    ws_rev.column_dimensions['B'].width = 12
    ws_rev.column_dimensions['C'].width = 14
    ws_rev.column_dimensions['D'].width = 22
    ws_rev.column_dimensions['E'].width = 12
    ws_rev.column_dimensions['F'].width = 12
    for i in range(NUM_QUARTERS):
        ws_rev.column_dimensions[get_column_letter(7+i)].width = 9
    for i in range(4):
        ws_rev.column_dimensions[get_column_letter(7+NUM_QUARTERS+i)].width = 10
    
    if not df_rev_sorted.empty:
        ws_rev.auto_filter.ref = f"B2:{get_column_letter(6+NUM_QUARTERS+4)}{len(df_rev_sorted)+2}"
        outdated_start = find_outdated_start(df_rev_sorted)
        if outdated_start:
            for r in range(outdated_start, len(df_rev_sorted) + 3):
                ws_rev.row_dimensions[r].hidden = True
    
    # ========================================
    # 3. Operating Income ÏãúÌä∏
    # ========================================
    ws_op = wb.create_sheet("Operating Income")
    
    op_headers = ['Ticker', 'Country', 'Industry', 'Latest_Date', 'Growth_Rate'] + QUARTER_LABELS + rolling_headers
    
    for col, header in enumerate(op_headers, start=2):
        cell = ws_op.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        if header in rolling_headers:
            cell.font = rolling_header_font
            cell.fill = rolling_header_fill
    
    for row_idx, row in df_op_sorted.iterrows():
        excel_row = row_idx + 3
        ws_op.cell(row=excel_row, column=2, value=row.get('Ticker', '')).font = pretendard
        ws_op.cell(row=excel_row, column=3, value=row.get('Country', '')).font = pretendard
        ws_op.cell(row=excel_row, column=4, value=row.get('Industry', '')).font = pretendard
        ws_op.cell(row=excel_row, column=5, value=row.get('Latest_Date', '')).font = pretendard
        
        growth_cell = ws_op.cell(row=excel_row, column=6)
        growth_val = row.get('Growth_Rate', 0)
        if pd.notna(growth_val):
            growth_cell.value = round(float(growth_val), 1)
            growth_cell.number_format = '0.0"%"'
        growth_cell.font = pretendard
        
        for i, q in enumerate(QUARTER_LABELS):
            cell = ws_op.cell(row=excel_row, column=7+i)
            val = row.get(q)
            if pd.notna(val) and val != '':
                cell.value = val
                cell.number_format = '#,##0'
            cell.font = pretendard
        
        # Î°§ÎßÅ ÏÑ±Ïû•Î•†
        if row_idx < len(op_rolling):
            for i, rv in enumerate(op_rolling[row_idx]):
                cell = ws_op.cell(row=excel_row, column=7+NUM_QUARTERS+i)
                if rv is not None:
                    cell.value = rv
                    cell.number_format = '0.0%'
                    if rv >= 0.1:
                        cell.fill = highlight_fill
                cell.font = pretendard
    
    for (r, c) in op_highlights:
        ws_op.cell(row=r, column=c).fill = highlight_fill
    
    ws_op.column_dimensions['A'].width = 3
    ws_op.column_dimensions['B'].width = 12
    ws_op.column_dimensions['C'].width = 14
    ws_op.column_dimensions['D'].width = 22
    ws_op.column_dimensions['E'].width = 12
    ws_op.column_dimensions['F'].width = 12
    for i in range(NUM_QUARTERS):
        ws_op.column_dimensions[get_column_letter(7+i)].width = 9
    for i in range(4):
        ws_op.column_dimensions[get_column_letter(7+NUM_QUARTERS+i)].width = 10
    
    if not df_op_sorted.empty:
        ws_op.auto_filter.ref = f"B2:{get_column_letter(6+NUM_QUARTERS+4)}{len(df_op_sorted)+2}"
        outdated_start_op = find_outdated_start(df_op_sorted)
        if outdated_start_op:
            for r in range(outdated_start_op, len(df_op_sorted) + 3):
                ws_op.row_dimensions[r].hidden = True
    
    # ========================================
    # 4. Îπ®Í∞ÑÏ§Ñ ÏãúÌä∏
    # ========================================
    ws_anal = wb.create_sheet("Îπ®Í∞ÑÏ§Ñ")
    
    ws_anal['B2'] = '(Îã®ÏúÑ: mil $)'
    ws_anal['B2'].font = Font(name='Pretendard', italic=True, color="666666", size=10)
    
    ws_anal['B4'] = 'STX'
    ws_anal['B4'].font = Font(name='Pretendard', bold=True, size=12)
    ws_anal['B4'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for i, q_label in enumerate(QUARTER_LABELS):
        col = 3 + i
        ws_anal.cell(row=4, column=col, value=6 + i)
        ws_anal.cell(row=4, column=col).font = Font(name='Pretendard', color="999999", size=9)
    
    for i, q_label in enumerate(QUARTER_LABELS):
        col = 3 + i
        cell = ws_anal.cell(row=6, column=col, value=q_label)
        cell.font = Font(name='Pretendard', bold=True, size=10)
        cell.alignment = Alignment(horizontal='center')
    
    rev_range = "Revenue!$B$2:$X$500"
    op_range = "'Operating Income'!$B$2:$X$500"
    
    # Revenue (Ìñâ 7)
    ws_anal['B7'] = 'Revenue'
    ws_anal['B7'].font = Font(name='Pretendard', bold=True)
    for i in range(NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        formula = f'=IFERROR(VLOOKUP($B$4,{rev_range},{col_letter}$4,FALSE),"-")'
        cell = ws_anal.cell(row=7, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '#,##0'
    
    # OP (Ìñâ 8)
    ws_anal['B8'] = 'OP'
    ws_anal['B8'].font = Font(name='Pretendard', bold=True)
    for i in range(NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        formula = f'=IFERROR(VLOOKUP($B$4,{op_range},{col_letter}$4,FALSE),"-")'
        cell = ws_anal.cell(row=8, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '#,##0'
    
    # Trailing 4Q OP avg (Ìñâ 9)
    ws_anal['B9'] = 'Trailing 4Q OP avg.'
    ws_anal['B9'].font = Font(name='Pretendard', bold=True)
    for i in range(NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        if i < 3:
            formula = f'=IFERROR(AVERAGE(C8:{col_letter}8),"-")'
        else:
            start_col = get_column_letter(col - 3)
            formula = f'=IFERROR(AVERAGE({start_col}8:{col_letter}8),"-")'
        cell = ws_anal.cell(row=9, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '#,##0.00'
    
    # Trailing 4Q OP sum (Ìñâ 10)
    ws_anal['B10'] = 'Trailing 4Q OP sum.'
    ws_anal['B10'].font = Font(name='Pretendard', bold=True)
    for i in range(NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        if i < 3:
            formula = f'=IFERROR(SUM(C8:{col_letter}8),"-")'
        else:
            start_col = get_column_letter(col - 3)
            formula = f'=IFERROR(SUM({start_col}8:{col_letter}8),"-")'
        cell = ws_anal.cell(row=10, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '#,##0'
    
    # Trailing OP Delta (Ìñâ 11)
    ws_anal['B11'] = 'Trailing OP Delta'
    ws_anal['B11'].font = Font(name='Pretendard', bold=True)
    for i in range(1, NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        prev_col_letter = get_column_letter(col - 1)
        formula = f'=IFERROR({col_letter}9/{prev_col_letter}9-1,"-")'
        cell = ws_anal.cell(row=11, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '0.0%'
    
    # OPM (Ìñâ 12)
    ws_anal['B12'] = 'OPM'
    ws_anal['B12'].font = Font(name='Pretendard', bold=True)
    for i in range(NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        formula = f'=IFERROR({col_letter}8/{col_letter}7,"-")'
        cell = ws_anal.cell(row=12, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '0.0%'
    
    # OP YoY (Ìñâ 13)
    ws_anal['B13'] = 'OP YoY'
    ws_anal['B13'].font = Font(name='Pretendard', bold=True)
    for i in range(4, NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        yoy_col_letter = get_column_letter(col - 4)
        formula = f'=IFERROR({col_letter}8/{yoy_col_letter}8-1,"-")'
        cell = ws_anal.cell(row=13, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '0.0%'
    
    # Îß§Ï∂ú YoY (Ìñâ 14)
    ws_anal['B14'] = 'Îß§Ï∂ú YoY'
    ws_anal['B14'].font = Font(name='Pretendard', bold=True)
    for i in range(4, NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        yoy_col_letter = get_column_letter(col - 4)
        formula = f'=IFERROR({col_letter}7/{yoy_col_letter}7-1,"-")'
        cell = ws_anal.cell(row=14, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '0.0%'
    
    # Ï°∞Í±¥Î∂ÄÏÑúÏãù
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    ws_anal.conditional_formatting.add(
        'D11:S11',
        CellIsRule(operator='greaterThanOrEqual', formula=['0.1'], fill=red_fill)
    )
    
    ws_anal.column_dimensions['A'].width = 3
    ws_anal.column_dimensions['B'].width = 18
    for i in range(NUM_QUARTERS):
        ws_anal.column_dimensions[get_column_letter(3 + i)].width = 9
    
    # Ï†ÄÏû•
    wb.save(output_file)
    print(f"\n‚úÖ ÏóëÏÖÄ Ï†ÄÏû• ÏôÑÎ£å: {output_file}")

# ==========================================
# Î©îÏù∏ Ïã§Ìñâ
# ==========================================
if __name__ == "__main__":
    print("="*50)
    print("Ìï¥Ïô∏ Ïã§Ï†Å ÌÅ¨Î°§ÎßÅ + ÏóëÏÖÄ ÏÉùÏÑ± ÌÜµÌï© Ïä§ÌÅ¨Î¶ΩÌä∏")
    print("="*50)
    print(f"‚ö†Ô∏è  ÏµúÏã† Ïã§Ï†Å Í∏∞Ï§Ä: {MIN_DATE[0]}ÎÖÑ {MIN_DATE[1]}Ïõî Ïù¥ÌõÑÎßå Ï†ïÏÉÅ Ï≤òÎ¶¨")
    
    # CSV ÏùΩÍ∏∞
    try:
        df_earnings = pd.read_csv(INPUT_CSV)
        print(f"\nüìÇ ÏûÖÎ†• ÌååÏùº: {INPUT_CSV}")
        print(f"   Ï¥ù {len(df_earnings)}Í∞ú Í∏∞ÏóÖ")
    except FileNotFoundError:
        print(f"‚ùå Ïò§Î•ò: '{INPUT_CSV}' ÌååÏùºÏùÑ Ï∞æÏùÑ Ïàò ÏóÜÏäµÎãàÎã§.")
        exit()
    
    # Ìã∞Ïª§ Ï∂îÏ∂ú
    ticker_col = next((col for col in df_earnings.columns if col.lower() == 'ticker'), None)
    if not ticker_col:
        print("‚ùå CSVÏóê 'Ticker' Ïª¨ÎüºÏù¥ ÏóÜÏäµÎãàÎã§.")
        exit()
    
    company_col = None
    for col in df_earnings.columns:
        if col.lower() in ['company', 'company name', 'name']:
            company_col = col
            break
    
    if company_col:
        ticker_list = list(zip(df_earnings[ticker_col].tolist(), df_earnings[company_col].tolist()))
    else:
        ticker_list = df_earnings[ticker_col].tolist()
    
    # ÌÅ¨Î°§ÎßÅ Ïã§Ìñâ
    print(f"\nüîÑ ÌÅ¨Î°§ÎßÅ ÏãúÏûë... (Î≥ëÎ†¨ {MAX_WORKERS}Í∞ú)")
    print("-"*50)
    
    success_data = []
    failed_data = []
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        results = list(executor.map(process_ticker, ticker_list))
    
    for res in results:
        if res['status'] == 'success':
            del res['status']
            success_data.append(res)
        else:
            failed_ticker = res['ticker'][0] if isinstance(res['ticker'], tuple) else res['ticker']
            failed_data.append({'Ticker': failed_ticker, 'Reason': res['reason']})
            print(f"[{failed_ticker}] ‚ùå {res['reason']}")
    
    print("-"*50)
    print(f"‚úÖ ÏÑ±Í≥µ: {len(success_data)}Í∞ú")
    print(f"‚ùå Ïã§Ìå®: {len(failed_data)}Í∞ú")
    
    # ÏÇ∞ÏóÖ Ï†ïÎ≥¥ Ï°∞Ìöå
    if success_data:
        print(f"\nüè≠ ÏÇ∞ÏóÖ Ï†ïÎ≥¥ Ï°∞Ìöå Ï§ë...")
        for data in success_data:
            data['industry'] = get_industry(data['ticker'])
            time.sleep(0.2)
        
        # ÏóëÏÖÄ ÏÉùÏÑ±
        print(f"\nüìä ÏóëÏÖÄ ÌååÏùº ÏÉùÏÑ± Ï§ë...")
        create_final_excel(success_data, df_earnings, OUTPUT_FILE)
    
    # Ïã§Ìå® Î™©Î°ù Ï†ÄÏû•
    if failed_data:
        df_failed = pd.DataFrame(failed_data)
        df_failed.to_excel(OUTPUT_FAILED, index=False)
        print(f"üìã Ïã§Ìå® Î™©Î°ù: {OUTPUT_FAILED}")
    
    print("\n" + "="*50)
    print("ÏôÑÎ£å!")
    print("="*50)
