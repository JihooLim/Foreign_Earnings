#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fiscal.ai -> (your template .xlsm) 자동 키인 스크립트
- 대상: Mercari (TSE-4385)
- 입력: Fiscal.ai API Key (환경변수 FISCAL_API_KEY 또는 --api-key)
- 출력: 템플릿(.xlsm) 내부 'Main FS(Re)' 시트에 분기 데이터 자동 입력
  -> '기업명' 시트는 기존 수식으로 자동 반영됨

요구사항:
  pip install requests openpyxl
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass
from typing import Dict, Any, List, Optional, Tuple

import requests
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

API_BASE = "https://api.fiscal.ai/v1"
DEFAULT_TICKER = "4385"
DEFAULT_MIC = "XTKS"  # Tokyo Stock Exchange MIC
DEFAULT_PERIODTYPE = "quarterly"

# 템플릿 'Main FS(Re)' 내 입력 대상 행 (필요시 여기만 수정하면 됨)
ROW_MAP = {
    "revenue": 849,          # 매출액(수익)
    "cogs": 881,             # 매출원가
    "gross_profit": 927,     # 매출총이익
    "operating_income": 992, # 영업이익
    "net_income": 1293,      # 당기순이익
}

# 템플릿에서 분기 라벨이 있는 위치(기업명 시트 2행, J열~)
TEMPLATE_QUARTER_ROW = 2
TEMPLATE_QUARTER_START_COL = "J"  # 1Q21 위치

# Summary에서 매출액 분기값을 끌어오는 Main FS(Re) 시작 열(= BP849가 1Q21에 대응)
MAINFS_QUARTER_START_COL = "BP"


@dataclass(frozen=True)
class Period:
    year: int
    quarter: int

    @staticmethod
    def from_label(label: str) -> Optional["Period"]:
        """
        템플릿 라벨 예:
          '1Q21', '4Q23', '4Q25E' (E는 무시)
        """
        if not isinstance(label, str):
            return None
        s = label.strip().upper().replace("’", "'")
        s = re.sub(r"[^0-9Q]", "", s)  # 숫자/Q만 남김 (E 제거)
        m = re.match(r"([1-4])Q(\d{2,4})$", s)
        if not m:
            return None
        q = int(m.group(1))
        yy = m.group(2)
        y = int(yy)
        if y < 100:
            # 00~99 -> 2000~2099로 가정
            y += 2000
        return Period(year=y, quarter=q)

    def key(self) -> str:
        return f"{self.year}Q{self.quarter}"


def _get_api_key(cli_key: Optional[str]) -> str:
    key = cli_key or os.environ.get("FISCAL_API_KEY")
    if not key:
        raise SystemExit(
            "Fiscal.ai API Key가 필요합니다.\n"
            "1) 환경변수 FISCAL_API_KEY 설정 또는\n"
            "2) 실행 시 --api-key YOUR_KEY 옵션을 주세요."
        )
    return key.strip()


def fetch_income_statement_standardized(
    api_key: str,
    ticker: str,
    mic: str,
    period_type: str = DEFAULT_PERIODTYPE,
) -> Dict[str, Any]:
    """
    docs: /v1/company/financials/income-statement/standardized
    """
    url = f"{API_BASE}/company/financials/income-statement/standardized"
    params = {
        "ticker": ticker,
        "micCode": mic,
        "periodType": period_type,
    }
    headers = {"X-Api-Key": api_key}
    r = requests.get(url, params=params, headers=headers, timeout=60)
    r.raise_for_status()
    return r.json()


def to_millions(value: float, unit: Optional[str]) -> float:
    """
    템플릿 Main FS(Re)는 '백만(=million) 단위' 입력을 전제로 설계된 경우가 많아,
    unit을 참고해 '백만 단위'로 정규화합니다.

    - unit이 없으면: "원단위"라고 가정하고 / 1e6
    - unit == "M": 이미 백만 단위
    - unit == "B": 10^3 백만
    - unit == "K": 10^-3 백만
    """
    if value is None:
        return float("nan")
    if unit is None:
        return float(value) / 1_000_000.0
    u = unit.upper().strip()
    if u == "M":
        return float(value)
    if u == "B":
        return float(value) * 1_000.0
    if u == "K":
        return float(value) / 1_000.0
    # unknown -> safest: treat as raw and scale to million
    return float(value) / 1_000_000.0


def parse_period_rows(payload: Dict[str, Any]) -> Dict[str, Dict[str, float]]:
    """
    standardized 응답을 템플릿 입력용 dict로 변환.

    반환 형태:
      {
        "2021Q1": {"revenue": 50000.0, "cogs": 12345.0, ...}  # 모두 '백만 단위'
      }
    """
    data = payload.get("data") or []
    out: Dict[str, Dict[str, float]] = {}

    # 가능한 키 후보(표준화 스키마가 버전에 따라 다를 수 있어 방어적으로 처리)
    KEY_CANDIDATES = {
        "revenue": ["revenue", "totalRevenue", "total_revenue"],
        "cogs": ["costOfRevenue", "costOfGoodsSold", "cogs", "cost_of_revenue"],
        "gross_profit": ["grossProfit", "gross_profit"],
        "operating_income": ["operatingIncome", "operating_profit", "operatingProfit", "operating_income"],
        "net_income": ["netIncome", "net_income", "netProfit", "net_profit"],
    }

    for row in data:
        cy = row.get("calendarYear")
        cq = row.get("calendarQuarter")
        if not (cy and cq):
            continue
        p = Period(int(cy), int(cq)).key()

        # 단위 정보가 row 단위로 있거나, 항목별로 있을 수 있어 둘 다 대응
        row_unit = row.get("unit")  # 있을 수도/없을 수도
        metrics = {}

        for out_key, candidates in KEY_CANDIDATES.items():
            val = None
            unit = row_unit
            for k in candidates:
                if k in row and row[k] is not None:
                    val = row[k]
                    break
                # item 객체로 들어오는 경우도 방어
                if k in row and isinstance(row[k], dict):
                    obj = row[k]
                    val = obj.get("value")
                    unit = obj.get("unit", unit)
                    break
            if val is None:
                continue
            metrics[out_key] = to_millions(val, unit)

        # 파생(없으면 계산)
        if "gross_profit" not in metrics and ("revenue" in metrics and "cogs" in metrics):
            metrics["gross_profit"] = metrics["revenue"] - metrics["cogs"]

        out[p] = metrics

    return out


def read_template_periods(wb: openpyxl.Workbook) -> List[Tuple[int, int, Period]]:
    """
    기업명 시트의 분기 라벨을 읽어서,
    (col_index, excel_col_letter, Period) 리스트로 반환
    """
    ws = wb["기업명"]
    start = column_index_from_string(TEMPLATE_QUARTER_START_COL)

    periods: List[Tuple[int, int, Period]] = []
    # 오른쪽으로 쭉 읽되, 빈칸이 3개 연속 나오면 종료
    empty_streak = 0
    for c in range(start, start + 120):
        label = ws.cell(TEMPLATE_QUARTER_ROW, c).value
        p = Period.from_label(label) if label is not None else None
        if p is None:
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue
        empty_streak = 0
        periods.append((c, c - start, p))  # (기업명 col index, offset, period)
    return periods


def write_to_mainfs(
    wb: openpyxl.Workbook,
    period_metrics: Dict[str, Dict[str, float]],
    template_periods: List[Tuple[int, int, Period]],
) -> int:
    """
    Main FS(Re) 시트의 (MAINFS_QUARTER_START_COL + offset) 열에 값 입력.
    Returns: 입력된 셀 개수
    """
    ws = wb["Main FS(Re)"]
    start_col = column_index_from_string(MAINFS_QUARTER_START_COL)

    written = 0
    for _, offset, period in template_periods:
        key = period.key()
        if key not in period_metrics:
            continue

        col = start_col + offset
        metrics = period_metrics[key]

        for k, row in ROW_MAP.items():
            if k not in metrics:
                continue
            value = metrics[k]
            if value != value:  # NaN
                continue
            cell = ws.cell(row=row, column=col)
            cell.value = float(value)
            written += 1

    return written


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True, help="기업분석틀 .xlsm 경로")
    ap.add_argument("--out", required=True, help="저장할 .xlsm 경로")
    ap.add_argument("--api-key", default=None, help="Fiscal.ai API Key (또는 환경변수 FISCAL_API_KEY)")
    ap.add_argument("--ticker", default=DEFAULT_TICKER)
    ap.add_argument("--mic", default=DEFAULT_MIC)
    ap.add_argument("--periodType", default=DEFAULT_PERIODTYPE, choices=["quarterly", "annual", "semi-annual", "ltm", "ytd", "latest"])
    args = ap.parse_args()

    api_key = _get_api_key(args.api_key)

    print("[1/4] 템플릿 로드...")
    wb = openpyxl.load_workbook(args.template, keep_vba=True)

    print("[2/4] 템플릿 분기 라벨 읽기...")
    tps = read_template_periods(wb)
    if not tps:
        raise SystemExit("기업명 시트에서 분기 라벨을 찾지 못했습니다. (2행 J열부터 확인)")

    print("[3/4] Fiscal.ai에서 재무 데이터 수집...")
    payload = fetch_income_statement_standardized(
        api_key=api_key, ticker=args.ticker, mic=args.mic, period_type=args.periodType
    )
    period_metrics = parse_period_rows(payload)

    print("[4/4] Main FS(Re) 시트에 자동 입력...")
    written = write_to_mainfs(wb, period_metrics, tps)

    wb.save(args.out)
    print(f"완료: {args.out}")
    print(f"입력된 셀 수: {written}")
    if written == 0:
        print("주의: 템플릿 분기(예: 1Q21)가 API의 calendarYear/calendarQuarter와 매칭이 안됐을 수 있습니다.")
        print(" - 템플릿 라벨이 '회계연도 기준'이면 Period.from_label 규칙을 바꿔야 합니다.")
        print(" - 또는 standardized 대신 as-reported를 사용해야 할 수 있습니다.")


if __name__ == "__main__":
    main()
