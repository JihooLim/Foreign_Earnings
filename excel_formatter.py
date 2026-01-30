from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import pandas as pd

# ==========================================
# 설정
# ==========================================
INPUT_CSV = "260127_Earnings.csv"  # 입력 CSV (기업리스트)
INPUT_XLSX = "quarterly_combined.xlsx"  # 크롤링 결과 (Revenue, Operating Income)
OUTPUT_FILE = "해외빨간줄_완성.xlsx"

NUM_QUARTERS = 17
QUARTER_LABELS = ['4Q21', '1Q22', '2Q22', '3Q22', '4Q22', '1Q23', '2Q23', '3Q23', '4Q23', '1Q24', '2Q24', '3Q24', '4Q24', '1Q25', '2Q25', '3Q25', '4Q25']

# ==========================================
# 정렬 함수
# ==========================================
def sort_dataframe(df):
    """Country(US→JP→나머지), Latest_Date(Dec'25 먼저), Growth_Rate 내림차순"""
    df = df.copy()
    
    def country_priority(c):
        c_str = str(c).strip().lower() if pd.notna(c) else ''
        if 'united states' in c_str or c_str == 'us':
            return 0
        elif 'japan' in c_str or c_str == 'jp':
            return 1
        else:
            return 2
    
    def is_dec25(date_str):
        if pd.isna(date_str):
            return False
        d = str(date_str).lower()
        return 'dec' in d and '25' in d
    
    df['_country_order'] = df['Country'].apply(country_priority)
    df['_is_dec25'] = df['Latest_Date'].apply(is_dec25)
    df['_growth'] = pd.to_numeric(df['Growth_Rate'], errors='coerce').fillna(0)
    
    df = df.sort_values(
        by=['_is_dec25', '_country_order', '_growth'],
        ascending=[False, True, False]
    )
    
    df = df.drop(columns=['_country_order', '_is_dec25', '_growth'])
    return df.reset_index(drop=True)

def find_non_dec25_start(df):
    """Dec'25 아닌 행의 시작 위치 찾기"""
    for i, row in df.iterrows():
        date_str = str(row.get('Latest_Date', '')).lower()
        if not ('dec' in date_str and '25' in date_str):
            return i + 3  # 헤더가 2행이므로
    return None

# ==========================================
# 롤링 성장률 계산
# ==========================================
def get_highlight_cells(df, quarter_cols, data_start_row=3, quarter_start_col=7):
    """롤링 4분기 성장률 10% 이상인 셀 좌표 반환"""
    highlights = []
    for row_idx, row in df.iterrows():
        for pos in range(4, len(quarter_cols)):
            try:
                recent_4q = [float(row[quarter_cols[pos-i]]) for i in range(4)]
                prev_4q = [float(row[quarter_cols[pos-1-i]]) for i in range(4)]
                recent_avg = sum(recent_4q) / 4
                prev_avg = sum(prev_4q) / 4
                if prev_avg != 0 and recent_avg / prev_avg >= 1.1:
                    excel_row = row_idx + data_start_row
                    excel_col = quarter_start_col + pos
                    highlights.append((excel_row, excel_col))
            except:
                continue
    return highlights

def calc_rolling_growth_for_quarters(df, quarter_cols):
    """1Q25, 2Q25, 3Q25, 4Q25 각각의 롤링 성장률 계산"""
    results = []
    target_indices = [13, 14, 15, 16]  # 1Q25, 2Q25, 3Q25, 4Q25
    
    for row_idx, row in df.iterrows():
        row_results = []
        for pos in target_indices:
            try:
                recent_4q = [float(row[quarter_cols[pos-i]]) for i in range(4)]
                prev_4q = [float(row[quarter_cols[pos-1-i]]) for i in range(4)]
                recent_avg = sum(recent_4q) / 4
                prev_avg = sum(prev_4q) / 4
                if prev_avg != 0:
                    growth = (recent_avg / prev_avg) - 1
                    row_results.append(growth)
                else:
                    row_results.append(None)
            except:
                row_results.append(None)
        results.append(row_results)
    return results

# ==========================================
# 메인 실행
# ==========================================
def create_final_excel(input_csv, input_xlsx, output_file):
    # 데이터 읽기
    df_earnings = pd.read_csv(input_csv)
    df_rev = pd.read_excel(input_xlsx, sheet_name='Revenue', header=1)
    df_op = pd.read_excel(input_xlsx, sheet_name='Operating Income', header=1)
    
    # Unnamed 컬럼 제거
    df_rev = df_rev.drop(columns=[c for c in df_rev.columns if 'Unnamed' in str(c)], errors='ignore')
    df_op = df_op.drop(columns=[c for c in df_op.columns if 'Unnamed' in str(c)], errors='ignore')
    
    # 정렬
    df_rev_sorted = sort_dataframe(df_rev)
    df_op_sorted = sort_dataframe(df_op)
    
    # 하이라이트 셀 계산
    rev_highlights = get_highlight_cells(df_rev_sorted, QUARTER_LABELS)
    op_highlights = get_highlight_cells(df_op_sorted, QUARTER_LABELS)
    op_rolling_growth = calc_rolling_growth_for_quarters(df_op_sorted, QUARTER_LABELS)
    
    # 워크북 생성
    wb = Workbook()
    
    # 스타일
    pretendard = Font(name='Pretendard', size=10)
    header_font = Font(name='Pretendard', size=10, bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    highlight_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    # ========================================
    # 1. 기업리스트 시트
    # ========================================
    ws_list = wb.active
    ws_list.title = "기업리스트"
    
    ws_list['B2'] = '실적 기업 정보'
    ws_list['B2'].font = Font(name='Pretendard', size=14, bold=True)
    
    list_headers = ['Ticker', 'Company', 'Date', 'Time', 'Quarter Ending', 'Market Cap (mil$)']
    for col, header in enumerate(list_headers, start=2):
        cell = ws_list.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, row in df_earnings.iterrows():
        excel_row = row_idx + 4
        ws_list.cell(row=excel_row, column=2, value=row.get('Ticker', '')).font = pretendard
        ws_list.cell(row=excel_row, column=3, value=row.get('Company', '')).font = pretendard
        ws_list.cell(row=excel_row, column=4, value=row.get('Date', '')).font = pretendard
        ws_list.cell(row=excel_row, column=5, value=row.get('Time', '')).font = pretendard
        ws_list.cell(row=excel_row, column=6, value=row.get('Quarter Ending', '')).font = pretendard
        
        market_cap = row.get(' Market Cap ', row.get('Market Cap', ''))
        if pd.notna(market_cap):
            try:
                mc_str = str(market_cap).replace(',', '').replace(' ', '').strip()
                mc_val = float(mc_str) / 1_000_000
                cell = ws_list.cell(row=excel_row, column=7, value=mc_val)
                cell.number_format = '#,##0'
            except:
                ws_list.cell(row=excel_row, column=7, value=market_cap)
        ws_list.cell(row=excel_row, column=7).font = pretendard
    
    ws_list.column_dimensions['A'].width = 3
    ws_list.column_dimensions['B'].width = 12
    ws_list.column_dimensions['C'].width = 45
    ws_list.column_dimensions['D'].width = 12
    ws_list.column_dimensions['E'].width = 12
    ws_list.column_dimensions['F'].width = 14
    ws_list.column_dimensions['G'].width = 16
    ws_list.auto_filter.ref = f"B3:G{len(df_earnings)+3}"
    
    # ========================================
    # 2. Revenue 시트
    # ========================================
    ws_rev = wb.create_sheet("Revenue")
    
    headers = ['Ticker', 'Country', 'Industry', 'Latest_Date', 'Growth_Rate'] + QUARTER_LABELS
    for col, header in enumerate(headers, start=2):
        cell = ws_rev.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, row in df_rev_sorted.iterrows():
        excel_row = row_idx + 3
        ws_rev.cell(row=excel_row, column=2, value=row.get('Ticker', '')).font = pretendard
        ws_rev.cell(row=excel_row, column=3, value=row.get('Country', '')).font = pretendard
        ws_rev.cell(row=excel_row, column=4, value=row.get('Industry', '')).font = pretendard
        ws_rev.cell(row=excel_row, column=5, value=row.get('Latest_Date', '')).font = pretendard
        
        growth_cell = ws_rev.cell(row=excel_row, column=6)
        growth_val = row.get('Growth_Rate', 0)
        if pd.notna(growth_val):
            growth_cell.value = round(float(growth_val), 1)
            growth_cell.number_format = '0.0"%"'
        growth_cell.font = pretendard
        
        for i, q in enumerate(QUARTER_LABELS):
            cell = ws_rev.cell(row=excel_row, column=7+i)
            val = row.get(q, '')
            if pd.notna(val):
                cell.value = val
                cell.number_format = '#,##0'
            cell.font = pretendard
    
    for (r, c) in rev_highlights:
        ws_rev.cell(row=r, column=c).fill = highlight_fill
    
    ws_rev.column_dimensions['A'].width = 3
    ws_rev.column_dimensions['B'].width = 12
    ws_rev.column_dimensions['C'].width = 14
    ws_rev.column_dimensions['D'].width = 22
    ws_rev.column_dimensions['E'].width = 18
    ws_rev.column_dimensions['F'].width = 12
    for i in range(NUM_QUARTERS):
        ws_rev.column_dimensions[get_column_letter(7+i)].width = 9
    
    ws_rev.auto_filter.ref = f"B2:{get_column_letter(6+NUM_QUARTERS)}{len(df_rev_sorted)+2}"
    
    non_dec25_start = find_non_dec25_start(df_rev_sorted)
    if non_dec25_start:
        for r in range(non_dec25_start, len(df_rev_sorted) + 3):
            ws_rev.row_dimensions[r].hidden = True
    
    # ========================================
    # 3. Operating Income 시트 (+ 롤링 성장률)
    # ========================================
    ws_op = wb.create_sheet("Operating Income")
    
    rolling_headers = ['1Q25 롤링', '2Q25 롤링', '3Q25 롤링', '4Q25 롤링']
    op_headers = ['Ticker', 'Country', 'Industry', 'Latest_Date', 'Growth_Rate'] + QUARTER_LABELS + rolling_headers
    
    for col, header in enumerate(op_headers, start=2):
        cell = ws_op.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        if header in rolling_headers:
            cell.font = Font(name='Pretendard', size=10, bold=True, color="CC0000")
            cell.fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
    
    for row_idx, row in df_op_sorted.iterrows():
        excel_row = row_idx + 3
        ws_op.cell(row=excel_row, column=2, value=row.get('Ticker', '')).font = pretendard
        ws_op.cell(row=excel_row, column=3, value=row.get('Country', '')).font = pretendard
        ws_op.cell(row=excel_row, column=4, value=row.get('Industry', '')).font = pretendard
        ws_op.cell(row=excel_row, column=5, value=row.get('Latest_Date', '')).font = pretendard
        
        growth_cell = ws_op.cell(row=excel_row, column=6)
        growth_val = row.get('Growth_Rate', 0)
        if pd.notna(growth_val):
            growth_cell.value = round(float(growth_val), 1)
            growth_cell.number_format = '0.0"%"'
        growth_cell.font = pretendard
        
        for i, q in enumerate(QUARTER_LABELS):
            cell = ws_op.cell(row=excel_row, column=7+i)
            val = row.get(q, '')
            if pd.notna(val):
                cell.value = val
                cell.number_format = '#,##0'
            cell.font = pretendard
        
        # 롤링 성장률 열
        rolling_vals = op_rolling_growth[row_idx]
        for i, rv in enumerate(rolling_vals):
            cell = ws_op.cell(row=excel_row, column=7+NUM_QUARTERS+i)
            if rv is not None:
                cell.value = rv
                cell.number_format = '0.0%'
                if rv >= 0.1:
                    cell.fill = highlight_fill
            cell.font = pretendard
    
    for (r, c) in op_highlights:
        ws_op.cell(row=r, column=c).fill = highlight_fill
    
    ws_op.column_dimensions['A'].width = 3
    ws_op.column_dimensions['B'].width = 12
    ws_op.column_dimensions['C'].width = 14
    ws_op.column_dimensions['D'].width = 22
    ws_op.column_dimensions['E'].width = 18
    ws_op.column_dimensions['F'].width = 12
    for i in range(NUM_QUARTERS):
        ws_op.column_dimensions[get_column_letter(7+i)].width = 9
    for i in range(4):
        ws_op.column_dimensions[get_column_letter(7+NUM_QUARTERS+i)].width = 10
    
    ws_op.auto_filter.ref = f"B2:{get_column_letter(6+NUM_QUARTERS+4)}{len(df_op_sorted)+2}"
    
    non_dec25_start_op = find_non_dec25_start(df_op_sorted)
    if non_dec25_start_op:
        for r in range(non_dec25_start_op, len(df_op_sorted) + 3):
            ws_op.row_dimensions[r].hidden = True
    
    # ========================================
    # 4. 빨간줄 시트 (조건부서식)
    # ========================================
    ws_anal = wb.create_sheet("빨간줄")
    
    ws_anal['B2'] = '(단위: mil $)'
    ws_anal['B2'].font = Font(name='Pretendard', italic=True, color="666666", size=10)
    
    ws_anal['B4'] = 'STX'
    ws_anal['B4'].font = Font(name='Pretendard', bold=True, size=12)
    ws_anal['B4'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # VLOOKUP 인덱스
    for i, q_label in enumerate(QUARTER_LABELS):
        col = 3 + i
        ws_anal.cell(row=4, column=col, value=6 + i)
        ws_anal.cell(row=4, column=col).font = Font(name='Pretendard', color="999999", size=9)
    
    # 분기 라벨
    for i, q_label in enumerate(QUARTER_LABELS):
        col = 3 + i
        cell = ws_anal.cell(row=6, column=col, value=q_label)
        cell.font = Font(name='Pretendard', bold=True, size=10)
        cell.alignment = Alignment(horizontal='center')
    
    rev_range = "Revenue!$B$2:$X$100"
    op_range = "'Operating Income'!$B$2:$X$100"
    
    # Revenue (행 7)
    ws_anal['B7'] = 'Revenue'
    ws_anal['B7'].font = Font(name='Pretendard', bold=True)
    for i in range(NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        formula = f'=IFERROR(VLOOKUP($B$4,{rev_range},{col_letter}$4,FALSE),"-")'
        cell = ws_anal.cell(row=7, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '#,##0'
    
    # OP (행 8)
    ws_anal['B8'] = 'OP'
    ws_anal['B8'].font = Font(name='Pretendard', bold=True)
    for i in range(NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        formula = f'=IFERROR(VLOOKUP($B$4,{op_range},{col_letter}$4,FALSE),"-")'
        cell = ws_anal.cell(row=8, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '#,##0'
    
    # Trailing 4Q OP avg (행 9)
    ws_anal['B9'] = 'Trailing 4Q OP avg.'
    ws_anal['B9'].font = Font(name='Pretendard', bold=True)
    for i in range(NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        if i < 3:
            formula = f'=IFERROR(AVERAGE(C8:{col_letter}8),"-")'
        else:
            start_col = get_column_letter(col - 3)
            formula = f'=IFERROR(AVERAGE({start_col}8:{col_letter}8),"-")'
        cell = ws_anal.cell(row=9, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '#,##0.00'
    
    # Trailing 4Q OP sum (행 10)
    ws_anal['B10'] = 'Trailing 4Q OP sum.'
    ws_anal['B10'].font = Font(name='Pretendard', bold=True)
    for i in range(NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        if i < 3:
            formula = f'=IFERROR(SUM(C8:{col_letter}8),"-")'
        else:
            start_col = get_column_letter(col - 3)
            formula = f'=IFERROR(SUM({start_col}8:{col_letter}8),"-")'
        cell = ws_anal.cell(row=10, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '#,##0'
    
    # Trailing OP Delta (행 11)
    ws_anal['B11'] = 'Trailing OP Delta'
    ws_anal['B11'].font = Font(name='Pretendard', bold=True)
    for i in range(1, NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        prev_col_letter = get_column_letter(col - 1)
        formula = f'=IFERROR({col_letter}9/{prev_col_letter}9-1,"-")'
        cell = ws_anal.cell(row=11, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '0.0%'
    
    # OPM (행 12)
    ws_anal['B12'] = 'OPM'
    ws_anal['B12'].font = Font(name='Pretendard', bold=True)
    for i in range(NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        formula = f'=IFERROR({col_letter}8/{col_letter}7,"-")'
        cell = ws_anal.cell(row=12, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '0.0%'
    
    # OP YoY (행 13)
    ws_anal['B13'] = 'OP YoY'
    ws_anal['B13'].font = Font(name='Pretendard', bold=True)
    for i in range(4, NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        yoy_col_letter = get_column_letter(col - 4)
        formula = f'=IFERROR({col_letter}8/{yoy_col_letter}8-1,"-")'
        cell = ws_anal.cell(row=13, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '0.0%'
    
    # 매출 YoY (행 14)
    ws_anal['B14'] = '매출 YoY'
    ws_anal['B14'].font = Font(name='Pretendard', bold=True)
    for i in range(4, NUM_QUARTERS):
        col = 3 + i
        col_letter = get_column_letter(col)
        yoy_col_letter = get_column_letter(col - 4)
        formula = f'=IFERROR({col_letter}7/{yoy_col_letter}7-1,"-")'
        cell = ws_anal.cell(row=14, column=col, value=formula)
        cell.font = Font(name='Pretendard')
        cell.number_format = '0.0%'
    
    # 조건부서식: Trailing OP Delta >= 10%
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    ws_anal.conditional_formatting.add(
        'D11:S11',
        CellIsRule(operator='greaterThanOrEqual', formula=['0.1'], fill=red_fill)
    )
    
    # 열 너비
    ws_anal.column_dimensions['A'].width = 3
    ws_anal.column_dimensions['B'].width = 18
    for i in range(NUM_QUARTERS):
        ws_anal.column_dimensions[get_column_letter(3 + i)].width = 9
    
    # 저장
    wb.save(output_file)
    print(f"✅ 완료: {output_file}")

if __name__ == "__main__":
    create_final_excel(INPUT_CSV, INPUT_XLSX, OUTPUT_FILE)
